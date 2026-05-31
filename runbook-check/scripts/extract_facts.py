#!/usr/bin/env python3
"""
Excel作業手順書から「構造を解釈せずに取れる事実」だけを抽出する。
どのセルがパラメータか・どれがコマンドか といった解釈は一切しない（サブエージェントの仕事）。

抽出する事実:
  - 全シートの非空セル（アドレス・解決値・関数文字列）
  - 全関数セルとその解決値・エラー有無・参照しているセル
  - エラー（#REF! 等）の一覧

出力: facts.json

使い方:
    python extract_facts.py runbook.xlsx -o facts.json [--no-recalc] [--max-cells 5000]

依存: openpyxl（必須）, LibreOffice（再計算。--no-recalc で既存キャッシュ値を使用）
"""
import argparse
import json
import os
import re
import shutil
import subprocess
import sys
import tempfile

try:
    from openpyxl import load_workbook
    from openpyxl.utils import get_column_letter, column_index_from_string
except ModuleNotFoundError:
    sys.stderr.write(
        "openpyxl が見つかりません。`pip install openpyxl` を実行してください"
        "（venv 推奨）。\n"
    )
    sys.exit(1)

ERROR_TOKENS = ("#REF!", "#DIV/0!", "#VALUE!", "#NAME?", "#N/A", "#NULL!", "#NUM!")
# 関数内のセル参照（例: パラメータ!B2, 'シート 名'!$B$2, B5）
SHEET_REF = re.compile(r"(?:'([^']+)'|([A-Za-z0-9_\u3040-\u30ff\u4e00-\u9fff]+))!\$?([A-Z]{1,3})\$?(\d+)")
LOCAL_REF = re.compile(r"(?<![A-Za-z0-9_!])\$?([A-Z]{1,3})\$?(\d+)(?![\w(])")


def _find_soffice():
    for name in ("libreoffice", "soffice"):
        p = shutil.which(name)
        if p:
            return p
    return None


def recalc(path):
    """関数の解決値を焼き直した**コピー**のパスを返す（元ファイルは変更しない）。

    再計算手段を順に試す:
      1) LibreOffice headless（`libreoffice`/`soffice`。`--convert-to xlsx` で読込→保存し再計算）
      2) claude.ai 同梱の recalc.py（あれば）
    どちらも使えなければ None を返し、呼び出し側は既存のキャッシュ値を使う。
    """
    td = tempfile.mkdtemp(prefix="runbook_recalc_")
    indir, outdir = os.path.join(td, "in"), os.path.join(td, "out")
    os.makedirs(indir)
    os.makedirs(outdir)
    work = os.path.join(indir, os.path.basename(path))
    shutil.copyfile(path, work)

    soffice = _find_soffice()
    if soffice:
        try:
            subprocess.run(
                [soffice, "--headless", "--calc", "--convert-to", "xlsx",
                 "--outdir", outdir, work],
                capture_output=True, timeout=120,
            )
            out = os.path.join(outdir, os.path.splitext(os.path.basename(path))[0] + ".xlsx")
            if os.path.exists(out):
                return out
        except Exception:
            pass

    script = "/mnt/skills/public/xlsx/scripts/recalc.py"
    if os.path.exists(script):
        try:
            subprocess.run([sys.executable, script, work, "60"], capture_output=True, timeout=120)
            return work
        except Exception:
            pass
    return None


def is_empty(v):
    return v is None or (isinstance(v, str) and v.strip() == "")


def is_error(v):
    return isinstance(v, str) and v.strip() in ERROR_TOKENS


def resolve_ref(wbv, ref):
    """'シート!セル' の解決値と、宛先が使用範囲内に存在するかを返す。
    戻り値: (value, exists)。シート不在や使用範囲外は exists=False。"""
    sheet, _, cell = ref.partition("!")
    m = re.match(r"^([A-Z]{1,3})(\d+)$", cell)
    if not m or sheet not in wbv.sheetnames:
        return None, False
    col = column_index_from_string(m.group(1))
    row = int(m.group(2))
    ws = wbv[sheet]
    exists = (row <= ws.max_row and col <= ws.max_column)
    val = ws.cell(row=row, column=col).value
    return val, exists


def mask_string_literals(formula):
    """Excel関数内の文字列リテラル "..."（内部の " は "" でエスケープ）を同じ長さの空白に置換する。
    description 等の文字列に含まれる "AS65001" "C100" 等をセル参照と誤認しないため。長さは保ち span を保つ。"""
    return re.sub(r'"(?:[^"]|"")*"', lambda m: " " * len(m.group(0)), formula or "")


def parse_refs(formula, current_sheet):
    """関数が参照するセルを 'シート!セル' 形式で返す（事実のみ）。文字列リテラル内は無視する。"""
    masked = mask_string_literals(formula)
    refs = []
    consumed = set()
    for m in SHEET_REF.finditer(masked):
        sheet = m.group(1) or m.group(2)
        refs.append(f"{sheet}!{m.group(3)}{m.group(4)}")
        consumed.add(m.span())
    # シート修飾のないローカル参照（同一シート）
    for m in LOCAL_REF.finditer(masked):
        # 既にシート参照の一部として消費した位置は除く
        if any(s <= m.start() < e for (s, e) in consumed):
            continue
        # 直前が "!" ならシート参照側で処理済み
        if m.start() > 0 and (masked[m.start() - 1] == "!"):
            continue
        refs.append(f"{current_sheet}!{m.group(1)}{m.group(2)}")
    # 重複除去（順序保持）
    seen, out = set(), []
    for r in refs:
        if r not in seen:
            seen.add(r); out.append(r)
    return out


def extract(path, do_recalc, max_cells):
    load_path = path
    recalc_applied = False
    if do_recalc:
        recalced = recalc(path)
        if recalced:
            load_path = recalced
            recalc_applied = True
    wbf = load_workbook(load_path, data_only=False)
    wbv = load_workbook(load_path, data_only=True)

    sheets = []
    formula_cells = []
    errors = []
    n_cells = n_formulas = 0
    truncated = False

    for name in wbf.sheetnames:
        wsf, wsv = wbf[name], wbv[name]
        cells = []
        for row in range(1, wsf.max_row + 1):
            for col in range(1, wsf.max_column + 1):
                fcell = wsf.cell(row=row, column=col)
                resolved = wsv.cell(row=row, column=col).value
                fval = fcell.value
                is_formula = isinstance(fval, str) and fval.startswith("=")
                if is_empty(resolved) and not is_formula:
                    continue
                if n_cells >= max_cells:
                    truncated = True
                    break
                addr = f"{get_column_letter(col)}{row}"
                entry = {"addr": addr, "value": resolved}
                if is_formula:
                    n_formulas += 1
                    refs = parse_refs(fval, name)
                    err = is_error(resolved) or any(t in fval for t in ERROR_TOKENS)
                    # 各参照先の解決値・空・宛先有無（誤参照判断の材料）
                    ref_details = []
                    empty_refs = []
                    dangling_refs = []
                    for r in refs:
                        rv, exists = resolve_ref(wbv, r)
                        empty = is_empty(rv)
                        ref_details.append({"ref": r, "value": rv, "empty": empty, "exists": exists})
                        if not exists:
                            dangling_refs.append(r)
                        elif empty:
                            empty_refs.append(r)
                    entry["formula"] = fval
                    if refs:
                        entry["refs"] = refs
                    if err:
                        entry["error"] = True
                    formula_cells.append({
                        "location": f"{name}!{addr}", "formula": fval,
                        "value": resolved, "error": bool(err),
                        "refs": refs, "ref_details": ref_details,
                        "empty_refs": empty_refs, "dangling_refs": dangling_refs,
                    })
                    if err:
                        errors.append({
                            "location": f"{name}!{addr}",
                            "error": resolved if is_error(resolved) else "#REF!",
                            "formula": fval,
                        })
                cells.append(entry)
                n_cells += 1
            if truncated:
                break
        sheets.append({
            "name": name,
            "max_row": wsf.max_row,
            "max_col": wsf.max_column,
            "cells": cells,
        })
        if truncated:
            break

    ref_issues = [
        {"location": fc["location"], "formula": fc["formula"], "value": fc["value"],
         "empty_refs": fc["empty_refs"], "dangling_refs": fc["dangling_refs"]}
        for fc in formula_cells if fc["empty_refs"] or fc["dangling_refs"]
    ]

    # 解決値(キャッシュ)が無い関数セル＝再計算が必要だったのにできなかった兆候
    unresolved = sum(1 for fc in formula_cells if fc["value"] is None)

    return {
        "workbook": os.path.basename(path),
        "sheet_names": wbf.sheetnames,
        "sheets": sheets,
        "formula_cells": formula_cells,
        "referenced_cells": sorted({r for fc in formula_cells for r in fc["refs"]}),
        "errors": errors,
        "ref_issues": ref_issues,
        "truncated": truncated,
        "recalc_applied": recalc_applied,
        "stats": {
            "sheets": len(wbf.sheetnames),
            "cells": n_cells,
            "formulas": n_formulas,
            "errors": len(errors),
            "ref_issues": len(ref_issues),
            "unresolved_formulas": unresolved,
        },
    }


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("workbook")
    ap.add_argument("-o", "--output", default="facts.json")
    ap.add_argument("--no-recalc", action="store_true", help="再計算せず既存キャッシュ値を使用")
    ap.add_argument("--max-cells", type=int, default=5000, help="抽出セル数の上限（巨大ファイル対策）")
    args = ap.parse_args()

    if not os.path.exists(args.workbook):
        print(f"ファイルが見つかりません: {args.workbook}", file=sys.stderr)
        sys.exit(1)

    result = extract(args.workbook, not args.no_recalc, args.max_cells)
    with open(args.output, "w", encoding="utf-8") as fp:
        json.dump(result, fp, ensure_ascii=False, indent=2)

    st = result["stats"]
    print(f"事実抽出完了: {args.output}")
    print(f"  シート {st['sheets']} / セル {st['cells']} / 関数 {st['formulas']} / エラー {st['errors']} / 参照問題(空・宛先なし) {st['ref_issues']}")
    if result["truncated"]:
        print(f"  ※ セル数が上限({args.max_cells})に達し打ち切り。--max-cells で拡張可。")
    if st.get("unresolved_formulas"):
        if result.get("recalc_applied"):
            print(f"  ※ 関数 {st['unresolved_formulas']} 個が再計算後も空。参照先が空の可能性（ref_issues を参照）。")
        else:
            print(f"  ※ 関数 {st['unresolved_formulas']} 個に解決値(キャッシュ)が無く、再計算もできませんでした。")
            print(f"     LibreOffice 導入（apt install libreoffice-calc 等）か、Excel で開いて再保存すると解決値が埋まります。")


if __name__ == "__main__":
    main()
