#!/usr/bin/env python3
"""
ディレクトリを走査して、作業手順書の候補となるExcelファイルを列挙する。
Phase 0 でユーザーに選択肢として提示するために使う。

使い方:
    python list_workbooks.py [DIR] [--recursive] [--max N]

出力JSON: 候補の配列（更新日時の新しい順）
  [{"index":1,"path":"...","name":"...","size_kb":..,"modified":"YYYY-MM-DD HH:MM","sheets":[...],"sheet_count":N}, ...]

除外: Excelロックファイル(~$...)・隠しファイル・本スキル同梱のsample。
依存: openpyxl（シート名取得。失敗してもファイル情報は返す）
"""
import argparse
import json
import os
import sys
from datetime import datetime

EXTS = (".xlsx", ".xlsm")


def sheet_names(path):
    try:
        from openpyxl import load_workbook
        wb = load_workbook(path, read_only=True, data_only=True)
        names = list(wb.sheetnames)
        wb.close()
        return names
    except Exception:
        return None


def scan(directory, recursive, max_n):
    cands = []
    if recursive:
        walker = ((r, f) for r, _, fs in os.walk(directory) for f in fs)
    else:
        walker = ((directory, f) for f in os.listdir(directory)
                  if os.path.isfile(os.path.join(directory, f)))
    for root, fname in walker:
        if fname.startswith("~$") or fname.startswith("."):
            continue
        if not fname.lower().endswith(EXTS):
            continue
        if fname == "sample_runbook.xlsx":
            continue
        path = os.path.join(root, fname)
        try:
            stat = os.stat(path)
        except OSError:
            continue
        cands.append({
            "path": os.path.abspath(path),
            "name": fname,
            "size_kb": round(stat.st_size / 1024, 1),
            "_mtime": stat.st_mtime,
            "modified": datetime.fromtimestamp(stat.st_mtime).strftime("%Y-%m-%d %H:%M"),
        })
    cands.sort(key=lambda c: c["_mtime"], reverse=True)
    cands = cands[:max_n]
    for i, c in enumerate(cands, 1):
        c["index"] = i
        names = sheet_names(c["path"])
        c["sheets"] = names or []
        c["sheet_count"] = len(names) if names is not None else None
        del c["_mtime"]
    return cands


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("directory", nargs="?", default=".", help="走査するディレクトリ（既定: カレント）")
    ap.add_argument("--recursive", action="store_true", help="サブディレクトリも走査")
    ap.add_argument("--max", type=int, default=30, help="最大候補数")
    args = ap.parse_args()

    if not os.path.isdir(args.directory):
        print(json.dumps({"error": f"ディレクトリが存在しません: {args.directory}", "candidates": []},
                         ensure_ascii=False))
        sys.exit(1)

    cands = scan(args.directory, args.recursive, args.max)
    print(json.dumps({"directory": os.path.abspath(args.directory),
                      "count": len(cands), "candidates": cands},
                     ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
