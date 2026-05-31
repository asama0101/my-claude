#!/usr/bin/env python3
"""
runbook-check の eval 用テスト手順書（.xlsx）を生成する。

ローカルに LibreOffice が無くてもスキルが関数の解決値を読めるよう、
openpyxl で関数を書いたうえで、各関数セルの**キャッシュ値（<v>）を XML に注入**して
「Excel/LibreOffice が保存した直後」相当のブックにする。

生成する3ブック（レイアウトと仕込みバグを変えてスキルの守備範囲を突く）:
  A) a-step-split-buggy.xlsx   STEP分割型・複数の仕込みバグ（#REF! / 空参照 / 不正値 / 切り戻し欠落 / 改版履歴未更新）
  B) b-single-sheet.xlsx       1シート密集型・別レイアウト（認証情報平文 / ハードコード / 事前確認欠落）
  C) c-clean-good.xlsx         パラメータ別シート参照型・ほぼ良好（誤検知の少なさ＋構成図の検証）

使い方:
    python make_fixtures.py [-o OUTDIR]   # 既定 OUTDIR=inputs/
"""
import argparse
import os
import shutil
import tempfile
import zipfile
from xml.etree import ElementTree as ET

from openpyxl import Workbook

MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
M = "{%s}" % MAIN_NS
R = "{%s}" % REL_NS


# ── 関数セルのキャッシュ値注入 ───────────────────────────────────────────────
def inject_cached(path, cached):
    """cached: {sheetname: {addr: (value, is_error)}} を関数セルの <v> として焼き込む。"""
    ET.register_namespace("", MAIN_NS)
    ET.register_namespace("r", REL_NS)
    tmp = tempfile.mkdtemp(prefix="fixture_")
    try:
        with zipfile.ZipFile(path) as z:
            z.extractall(tmp)
        # シート名 → worksheet xml ファイル
        wb = ET.parse(os.path.join(tmp, "xl", "workbook.xml"))
        name_to_rid = {s.get("name"): s.get(R + "id")
                       for s in wb.find(M + "sheets").findall(M + "sheet")}
        rels = ET.parse(os.path.join(tmp, "xl", "_rels", "workbook.xml.rels"))
        rid_to_target = {r.get("Id"): r.get("Target") for r in rels.getroot()}

        for sheetname, addrs in cached.items():
            target = rid_to_target[name_to_rid[sheetname]]
            # Target は "/xl/worksheets/sheet1.xml"(絶対) か "worksheets/sheet1.xml"(xl相対)
            if target.startswith("/"):
                wspath = os.path.join(tmp, target.lstrip("/"))
            else:
                wspath = os.path.join(tmp, "xl", target)
            tree = ET.parse(wspath)
            cmap = {c.get("r"): c for c in tree.getroot().iter(M + "c")}
            for addr, (val, is_err) in addrs.items():
                c = cmap.get(addr)
                if c is None:
                    continue
                for v in c.findall(M + "v"):
                    c.remove(v)
                v = ET.SubElement(c, M + "v")  # <f> の後ろに付く（スキーマ順OK）
                c.set("t", "e" if is_err else "str")
                v.text = str(val)
            tree.write(wspath, xml_declaration=True, encoding="UTF-8")

        os.remove(path)
        with zipfile.ZipFile(path, "w", zipfile.ZIP_DEFLATED) as z:
            for root, _, files in os.walk(tmp):
                for f in files:
                    full = os.path.join(root, f)
                    z.write(full, os.path.relpath(full, tmp))
    finally:
        shutil.rmtree(tmp, ignore_errors=True)


class Builder:
    """関数式と解決値を一緒に組み立て、注入用 cached を貯める補助。"""

    def __init__(self):
        self.wb = Workbook()
        self.wb.remove(self.wb.active)
        self.cached = {}

    def sheet(self, name):
        ws = self.wb.create_sheet(name)
        self.cached.setdefault(name, {})
        return ws

    def lit(self, text):
        return ("lit", text)

    def ref(self, ref, value):
        return ("ref", ref, value)

    def formula(self, ws, addr, *parts, error=False):
        """parts を & 連結した関数式をセルへ。解決値を cached に記録。
        error=True なら #REF! を式に含め、値もエラーにする。"""
        fp, rp = [], []
        for p in parts:
            if p[0] == "lit":
                fp.append('"%s"' % p[1].replace('"', '""'))
                rp.append(p[1])
            else:  # ref
                fp.append(p[1])
                rp.append("" if p[2] is None else str(p[2]))
        if error:
            formula = '=CONCATENATE(' + ",".join(fp + ["#REF!"]) + ')'
            ws[addr] = formula
            self.cached[ws.title][addr] = ("#REF!", True)
        else:
            ws[addr] = "=" + "&".join(fp)
            self.cached[ws.title][addr] = ("".join(rp), False)

    def save(self, path):
        self.wb.save(path)
        inject_cached(path, self.cached)


# ── A) STEP分割型・仕込みバグ多数 ───────────────────────────────────────────
def build_a(path):
    b = Builder()

    rev = b.sheet("改版履歴")
    rev.append(["版数", "日付", "変更内容", "担当"])
    rev.append(["v1.0", "2026-03-01", "初版", "A"])
    rev.append(["v1.1", "2026-04-10", "IF設定追加", "B"])
    # → 今回(v1.2想定)の行が無い＝改版履歴 未更新

    p = b.sheet("パラメータ")
    p.append(["パラメータ名", "値", "型", "許容範囲"])
    p.append(["ホスト名", "GTE-02", "hostname", ""])            # B2
    p.append(["管理IP", "192.0.2.300", "ip", ""])               # B3 ← 不正IP(300)
    p.append(["サブネットマスク", 24, "mask", ""])               # B4
    p.append(["VLAN-ID", 5000, "vlan", "1-4094"])               # B5 ← 範囲外(>4094)
    p.append(["IF名", "ge-0/0/5", "interface", ""])             # B6
    p.append(["description", "", "text", ""])                   # B7 ← 空（空参照の元）

    s1 = b.sheet("STEP1")
    s1.append(["STEP", "操作", "コマンド"])
    b.formula(s1, "C2", b.lit("set system host-name "), b.ref("パラメータ!B2", "GTE-02"))
    s1["A2"], s1["B2"] = 1, "ホスト名設定"
    # B7(description)が空 → 空参照
    b.formula(s1, "C3", b.lit("set interfaces "), b.ref("パラメータ!B6", "ge-0/0/5"),
              b.lit(" description "), b.ref("パラメータ!B7", ""))
    s1["A3"], s1["B3"] = 2, "IF description設定"
    b.formula(s1, "C4", b.lit("set vlans VLAN"), b.ref("パラメータ!B5", 5000),
              b.lit(" vlan-id "), b.ref("パラメータ!B5", 5000))
    s1["A4"], s1["B4"] = 3, "VLAN割当"

    s2 = b.sheet("STEP2")
    s2.append(["STEP", "操作", "コマンド"])
    b.formula(s2, "C2", b.lit("set interfaces "), b.ref("パラメータ!B6", "ge-0/0/5"),
              b.lit(" unit 0 family inet address "), b.ref("パラメータ!B3", "192.0.2.300"),
              b.lit("/"), b.ref("パラメータ!B4", 24))
    s2["A2"], s2["B2"] = 1, "管理IP設定"
    b.formula(s2, "C3", b.lit("delete protocols ospf area 0 interface "), error=True)  # #REF!
    s2["A3"], s2["B3"] = 2, "旧OSPF設定削除"

    s3 = b.sheet("STEP3")
    s3.append(["STEP", "操作", "コマンド"])
    s3["A2"], s3["B2"], s3["C2"] = 1, "保存", "commit and-quit"
    # 切り戻し手順・事前バックアップ・正常性確認・影響範囲の記載なし（レビュアーが検出）

    b.save(path)


# ── B) 1シート密集型・別レイアウト・セキュリティ ───────────────────────────
def build_b(path):
    b = Builder()
    w = b.sheet("作業手順")
    w["A1"] = "■ パラメータ"
    w["A2"], w["B2"] = "ホスト名", "EDGE-FW-01"
    w["A3"], w["B3"] = "ループバックIP", "10.0.0.9"
    w["A4"], w["B4"] = "SNMPコミュニティ", "public"          # 平文コミュニティ
    w["A5"], w["B5"] = "rootパスワード", "P@ssw0rd123"        # 平文パスワード直書き

    w["A7"] = "■ 投入手順"
    w["A8"], w["B8"] = "1 ホスト名", None
    b.formula(w, "C8", b.lit("set system host-name "), b.ref("B2", "EDGE-FW-01"))
    w["A9"] = "2 ループバック"
    b.formula(w, "C9", b.lit("set interfaces lo0 unit 0 family inet address "),
              b.ref("B3", "10.0.0.9"), b.lit("/32"))
    w["A10"] = "3 SNMP"
    b.formula(w, "C10", b.lit("set snmp community "), b.ref("B4", "public"),
              b.lit(" authorization read-only"))
    w["A11"] = "4 root認証"
    b.formula(w, "C11", b.lit("set system root-authentication plain-text-password "),
              b.ref("B5", "P@ssw0rd123"))
    w["A12"], w["C12"] = "5 NTP", None
    b.formula(w, "C12", b.lit("set system ntp server 192.0.2.123"))  # ハードコード(参照なし)
    w["A13"], w["B13"] = "6 接続", "Telnetで対象機器へログインして投入する"  # 平文プロトコル
    w["A14"] = "7 反映"
    b.formula(w, "C14", b.lit("commit"))
    # 事前確認・バックアップ・切り戻し・影響範囲の記載なし

    b.save(path)


# ── C) パラメータ別シート参照型・ほぼ良好（誤検知チェック） ─────────────────
def build_c(path):
    b = Builder()

    rev = b.sheet("改版履歴")
    rev.append(["版数", "日付", "変更内容", "担当"])
    rev.append(["v1.0", "2026-03-01", "初版", "A"])
    rev.append(["v1.1", "2026-04-15", "確認手順追記", "B"])
    rev.append(["v1.2", "2026-05-30", "アクセスVLAN(300)追加", "C"])  # 今回分を更新済み

    o = b.sheet("目的・前提")
    o.append(["項目", "内容"])
    o.append(["目的", "GTE-03 に新規アクセスVLAN 300 を追加し、ge-0/0/7 をアクセスポートとして収容する"])
    o.append(["対象機器", "GTE-03（管理IP 192.0.2.23）"])
    o.append(["接続経路", "踏み台 bastion01 → SSH"])
    o.append(["メンテ枠", "2026-06-01 01:00-03:00"])
    o.append(["影響範囲", "ge-0/0/7 配下の端末のみ。既存VLAN・他ポートへの影響なし。瞬断なし"])
    o.append(["事前バックアップ", "show configuration | save backup-20260601.conf を取得済みであること"])

    p = b.sheet("パラメータ")
    p.append(["パラメータ名", "値", "型", "許容範囲"])
    p.append(["ホスト名", "GTE-03", "hostname", ""])           # B2
    p.append(["IF名", "ge-0/0/7", "interface", ""])           # B3
    p.append(["VLAN-ID", 300, "vlan", "1-4094"])              # B4

    s1 = b.sheet("STEP1_作業")
    s1.append(["STEP", "操作", "コマンド"])
    b.formula(s1, "C2", b.lit("set vlans v"), b.ref("パラメータ!B4", 300),
              b.lit(" vlan-id "), b.ref("パラメータ!B4", 300))
    s1["A2"], s1["B2"] = 1, "VLAN作成"
    b.formula(s1, "C3", b.lit("set interfaces "), b.ref("パラメータ!B3", "ge-0/0/7"),
              b.lit(" unit 0 family ethernet-switching interface-mode access vlan members v"),
              b.ref("パラメータ!B4", 300))
    s1["A3"], s1["B3"] = 2, "アクセスポート収容"

    s2 = b.sheet("STEP2_確認")
    s2.append(["STEP", "種別", "確認内容", "期待値"])
    s2.append([1, "作業前", "show vlans | match 300", "v300 が存在しないこと"])
    s2.append([2, "作業前", "show ethernet-switching interface ge-0/0/7", "現行の所属VLANを記録（比較基準）"])
    s2.append([3, "作業後", "show vlans v300", "ge-0/0/7.0 が members に表示される"])
    s2.append([4, "作業後", "show ethernet-switching interface ge-0/0/7", "VLAN v300・mode access が表示される"])

    rb = b.sheet("切り戻し")
    rb.append(["STEP", "操作", "コマンド/内容"])
    rb.append([1, "判断基準", "STEP2 の作業後確認でVLAN未収容なら切り戻す。判断者: 当番リーダー"])
    rb.append([2, "切り戻し", "rollback 0 を投入（未commit前提）。commit 済みなら下記 delete を投入"])
    rb.append([3, "切り戻し", "delete interfaces ge-0/0/7 unit 0 family ethernet-switching vlan members v300 ; delete vlans v300"])
    rb.append([4, "確認", "show vlans に v300 が無いこと・ge-0/0/7 が元の所属に戻ったこと"])
    rb.append([5, "時間", "メンテ枠内（〜03:00）に完了できること"])

    b.save(path)


# ── D) 複数機器（2台）・VLAN→IP階層・BGP・非Config混入 ─────────────────────
def build_d(path):
    b = Builder()

    rev = b.sheet("改版履歴")
    rev.append(["版数", "日付", "変更内容", "担当"])
    rev.append(["v1.0", "2026-05-01", "初版", "A"])
    rev.append(["v1.1", "2026-05-28", "CORE-EDGE間 eBGP 新設", "B"])

    p = b.sheet("パラメータ")
    p.append(["機器", "パラメータ名", "値", "型", "許容範囲"])
    p.append(["CORE-01", "IF名", "ge-0/0/1", "interface", ""])     # D2
    p.append(["CORE-01", "VLAN-ID", 100, "vlan", "1-4094"])        # D3
    p.append(["CORE-01", "IRB-IP", "10.0.0.1", "ip", ""])         # D4
    p.append(["CORE-01", "マスク", 30, "mask", ""])               # D5
    p.append(["CORE-01", "対向IP", "10.0.0.2", "ip", ""])         # D6
    p.append(["CORE-01", "自AS", 65000, "int", ""])              # D7
    p.append(["CORE-01", "対向AS", 65001, "int", ""])            # D8
    p.append(["EDGE-02", "IF名", "ge-0/0/1", "interface", ""])    # D9
    p.append(["EDGE-02", "VLAN-ID", 100, "vlan", "1-4094"])       # D10
    p.append(["EDGE-02", "IRB-IP", "10.0.0.2", "ip", ""])        # D11
    p.append(["EDGE-02", "マスク", 30, "mask", ""])              # D12
    p.append(["EDGE-02", "対向IP", "10.0.0.1", "ip", ""])        # D13
    p.append(["EDGE-02", "自AS", 65001, "int", ""])             # D14
    p.append(["EDGE-02", "対向AS", 65000, "int", ""])           # D15

    def device_sheet(name, dev, ifc, vlan, irb, mask, peer, asn, peer_as, with_commit):
        s = b.sheet(name)
        s.append(["STEP", "操作", "コマンド"])
        b.formula(s, "C2", b.lit("set interfaces "), b.ref("パラメータ!" + ifc[0], ifc[1]),
                  b.lit(f' description "to {peer[2]}"'))
        s["A2"], s["B2"] = 1, "IF description"
        b.formula(s, "C3", b.lit("set vlans v"), b.ref("パラメータ!" + vlan[0], vlan[1]),
                  b.lit(" vlan-id "), b.ref("パラメータ!" + vlan[0], vlan[1]))
        s["A3"], s["B3"] = 2, "VLAN作成"
        b.formula(s, "C4", b.lit("set vlans v"), b.ref("パラメータ!" + vlan[0], vlan[1]),
                  b.lit(" l3-interface irb."), b.ref("パラメータ!" + vlan[0], vlan[1]))
        s["A4"], s["B4"] = 3, "VLANにIRB割当"
        b.formula(s, "C5", b.lit("set interfaces "), b.ref("パラメータ!" + ifc[0], ifc[1]),
                  b.lit(" unit 0 family ethernet-switching interface-mode access vlan members v"),
                  b.ref("パラメータ!" + vlan[0], vlan[1]))
        s["A5"], s["B5"] = 4, "ポート収容"
        b.formula(s, "C6", b.lit("set interfaces irb unit "), b.ref("パラメータ!" + vlan[0], vlan[1]),
                  b.lit(" family inet address "), b.ref("パラメータ!" + irb[0], irb[1]),
                  b.lit("/"), b.ref("パラメータ!" + mask[0], mask[1]))
        s["A6"], s["B6"] = 5, "IRBにL3アドレス"
        b.formula(s, "C7", b.lit("set protocols bgp group ext neighbor "), b.ref("パラメータ!" + peer[0], peer[1]),
                  b.lit(" peer-as "), b.ref("パラメータ!" + peer_as[0], peer_as[1]))
        s["A7"], s["B7"] = 6, "eBGP neighbor"
        b.formula(s, "C8", b.lit("set routing-options autonomous-system "), b.ref("パラメータ!" + asn[0], asn[1]))
        s["A8"], s["B8"] = 7, "自AS設定"
        row = 9
        if with_commit:
            b.formula(s, f"C{row}", b.lit("commit"))
            s[f"A{row}"], s[f"B{row}"] = 8, "保存"
            row += 1
        # 非Config（地の文）— 投入Configから除外されるべき
        s[f"A{row}"], s[f"B{row}"] = row - 7, "接続"
        s[f"C{row}"] = f"Telnetで {dev} にログインして上記を投入する"

    # 値は C 列（A=機器,B=パラメータ名,C=値,D=型,E=許容範囲）。関数は C 列を参照する。
    # CORE-01: commit あり / EDGE-02: commit 無し（保存漏れ＝operability指摘の仕込み）
    device_sheet("CORE-01_設定", "CORE-01",
                 ("C2", "ge-0/0/1"), ("C3", 100), ("C4", "10.0.0.1"), ("C5", 30),
                 ("C6", "10.0.0.2", "EDGE-02 AS65001"), ("C7", 65000), ("C8", 65001), with_commit=True)
    device_sheet("EDGE-02_設定", "EDGE-02",
                 ("C9", "ge-0/0/1"), ("C10", 100), ("C11", "10.0.0.2"), ("C12", 30),
                 ("C13", "10.0.0.1", "CORE-01 AS65000"), ("C14", 65001), ("C15", 65000), with_commit=False)

    chk = b.sheet("確認")
    chk.append(["STEP", "種別", "確認内容", "期待値"])
    chk.append([1, "作業前", "show bgp summary（両機）", "対象 neighbor が未確立 or 未設定であること"])
    chk.append([2, "作業後", "show bgp summary（両機）", "10.0.0.1↔10.0.0.2 が Establ、受信経路数 >0"])
    chk.append([3, "作業後", "ping 10.0.0.2 source 10.0.0.1（CORE-01から）", "5/5 応答・RTT 平常±1ms"])

    rb = b.sheet("切り戻し")
    rb.append(["STEP", "操作", "コマンド/内容"])
    rb.append([1, "判断基準", "作業後の show bgp summary が Establ にならなければ切り戻す。判断者: 当番リーダー"])
    rb.append([2, "切り戻し", "両機で rollback 0（未commit前提）。commit済みは下記 delete を投入"])
    rb.append([3, "切り戻し", "delete protocols bgp group ext ; delete interfaces irb unit 100 ; delete vlans v100"])
    rb.append([4, "確認", "両機で show bgp summary に当該 neighbor が無いこと・既存経路に影響が無いこと"])

    b.save(path)


# ── E) 物理(FDF/TIE)＋論理(BGP/static)・既存と追加の混在（構成図HTML検証用）──
def build_e(path):
    b = Builder()

    rev = b.sheet("改版履歴")
    rev.append(["版数", "日付", "変更内容", "担当"])
    rev.append(["v1.0", "2026-05-10", "初版", "A"])
    rev.append(["v1.1", "2026-05-29", "CORE-EDGE間 区間A eBGP・static新設", "B"])

    inp = b.sheet("インプット情報")
    inp.append(["項目", "内容"])
    inp.append(["物理経路(新設区間)", "CORE-01 ge-0/0/2 ─ FDF #12 ─ TIE T-07 ─ FDF #34 ─ EDGE-02 ge-0/0/2"])
    inp.append(["既存IF", "CORE-01 ge-0/0/0 (description: uplink-A) 203.0.113.1/30"])
    inp.append(["既存IF", "EDGE-02 ge-0/0/0 (description: uplink-B) 203.0.113.5/30"])
    inp.append(["備考", "FDF=光配線盤・TIE=局間タイ回線（いずれも既設）。今回 ge-0/0/2・BGP・static を新規追加"])

    p = b.sheet("パラメータ")
    p.append(["機器", "パラメータ名", "値", "型", "許容範囲"])
    p.append(["CORE-01", "IF名", "ge-0/0/2", "interface", ""])        # C2
    p.append(["CORE-01", "自IP", "10.0.0.1", "ip", ""])              # C3
    p.append(["CORE-01", "マスク", 30, "mask", ""])                  # C4
    p.append(["CORE-01", "対向IP", "10.0.0.2", "ip", ""])            # C5
    p.append(["CORE-01", "自AS", 65000, "int", ""])                # C6
    p.append(["CORE-01", "対向AS", 65001, "int", ""])              # C7
    p.append(["CORE-01", "staticプレフィクス", "172.16.0.0/16", "cidr", ""])  # C8
    p.append(["CORE-01", "staticネクストホップ", "10.0.0.2", "ip", ""])       # C9
    p.append(["EDGE-02", "IF名", "ge-0/0/2", "interface", ""])        # C10
    p.append(["EDGE-02", "自IP", "10.0.0.2", "ip", ""])              # C11
    p.append(["EDGE-02", "マスク", 30, "mask", ""])                  # C12
    p.append(["EDGE-02", "対向IP", "10.0.0.1", "ip", ""])            # C13
    p.append(["EDGE-02", "自AS", 65001, "int", ""])                # C14
    p.append(["EDGE-02", "対向AS", 65000, "int", ""])              # C15
    p.append(["EDGE-02", "staticプレフィクス", "172.16.0.0/16", "cidr", ""])  # C16
    p.append(["EDGE-02", "staticネクストホップ", "10.0.0.1", "ip", ""])       # C17

    def dev_sheet(name, peer, ifc, ip, mask, nbr, asn, pas, pfx, nh):
        s = b.sheet(name)
        s.append(["STEP", "操作", "コマンド"])
        b.formula(s, "C2", b.lit("set interfaces "), b.ref("パラメータ!" + ifc[0], ifc[1]),
                  b.lit(f' description "to {peer} via FDF#12/TIE07/FDF#34"'))
        s["A2"], s["B2"] = 1, "IF description"
        b.formula(s, "C3", b.lit("set interfaces "), b.ref("パラメータ!" + ifc[0], ifc[1]),
                  b.lit(" unit 0 family inet address "), b.ref("パラメータ!" + ip[0], ip[1]),
                  b.lit("/"), b.ref("パラメータ!" + mask[0], mask[1]))
        s["A3"], s["B3"] = 2, "L3アドレス"
        b.formula(s, "C4", b.lit("set protocols bgp group ext type external"))
        s["A4"], s["B4"] = 3, "BGP group"
        b.formula(s, "C5", b.lit("set protocols bgp group ext neighbor "), b.ref("パラメータ!" + nbr[0], nbr[1]),
                  b.lit(" peer-as "), b.ref("パラメータ!" + pas[0], pas[1]))
        s["A5"], s["B5"] = 4, "BGP neighbor"
        b.formula(s, "C6", b.lit("set routing-options autonomous-system "), b.ref("パラメータ!" + asn[0], asn[1]))
        s["A6"], s["B6"] = 5, "自AS"
        b.formula(s, "C7", b.lit("set routing-options static route "), b.ref("パラメータ!" + pfx[0], pfx[1]),
                  b.lit(" next-hop "), b.ref("パラメータ!" + nh[0], nh[1]))
        s["A7"], s["B7"] = 6, "static route"
        b.formula(s, "C8", b.lit("commit"))
        s["A8"], s["B8"] = 7, "保存"
        s["A9"], s["B9"], s["C9"] = 8, "接続", f"Telnetで {name.split('_')[0]} にログインして上記を投入する"

    dev_sheet("CORE-01_設定", "EDGE-02",
              ("C2", "ge-0/0/2"), ("C3", "10.0.0.1"), ("C4", 30), ("C5", "10.0.0.2"),
              ("C6", 65000), ("C7", 65001), ("C8", "172.16.0.0/16"), ("C9", "10.0.0.2"))
    dev_sheet("EDGE-02_設定", "CORE-01",
              ("C10", "ge-0/0/2"), ("C11", "10.0.0.2"), ("C12", 30), ("C13", "10.0.0.1"),
              ("C14", 65001), ("C15", 65000), ("C16", "172.16.0.0/16"), ("C17", "10.0.0.1"))

    chk = b.sheet("確認")
    chk.append(["STEP", "種別", "確認内容", "期待値"])
    chk.append([1, "作業前", "show bgp summary（両機）", "対象 neighbor が未確立であること"])
    chk.append([2, "作業後", "show bgp summary（両機）", "10.0.0.1↔10.0.0.2 が Establ・受信経路>0"])
    chk.append([3, "作業後", "show route 172.16.0.0/16", "static が有効・next-hop が対向"])

    rb = b.sheet("切り戻し")
    rb.append(["STEP", "操作", "コマンド/内容"])
    rb.append([1, "判断基準", "作業後の bgp が Establ しなければ切り戻す。判断者: 当番リーダー"])
    rb.append([2, "切り戻し", "両機 rollback 0（未commit前提）。commit済みは下記 delete"])
    rb.append([3, "切り戻し", "delete protocols bgp group ext ; delete routing-options static route 172.16.0.0/16 ; delete interfaces ge-0/0/2 unit 0 family inet"])
    rb.append([4, "確認", "両機 show bgp summary に当該 neighbor が無いこと・既存経路に影響なし"])

    b.save(path)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("-o", "--outdir", default=os.path.join(os.path.dirname(__file__), "inputs"))
    args = ap.parse_args()
    os.makedirs(args.outdir, exist_ok=True)
    targets = {
        "a-step-split-buggy.xlsx": build_a,
        "b-single-sheet.xlsx": build_b,
        "c-clean-good.xlsx": build_c,
        "d-multi-device.xlsx": build_d,
        "e-physical-topology.xlsx": build_e,
    }
    for name, fn in targets.items():
        path = os.path.join(args.outdir, name)
        fn(path)
        print(f"生成: {path}")


if __name__ == "__main__":
    main()
