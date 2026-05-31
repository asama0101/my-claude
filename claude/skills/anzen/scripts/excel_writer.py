"""既存 xlsx のセルを上書きし、別パスに保存する（non-destructive）。"""
import argparse
import json
from pathlib import Path

import openpyxl


def fill_workbook(input_path: str, mapping: dict, output_path: str) -> None:
    """
    input_path の xlsx を読み込み、mapping で指定したセルを書き換えて
    output_path に保存する。input_path は変更しない。

    mapping: {"シート名": {"セル座標": 値, ...}, ...}

    Raises:
        ValueError: output_path が input_path と同一の場合（元ファイル保護）
        KeyError: mapping に存在しないシート名が含まれる場合
    """
    if Path(input_path).resolve() == Path(output_path).resolve():
        raise ValueError(f"output_path must differ from input_path: {input_path!r}")

    wb = openpyxl.load_workbook(input_path)

    # 書き込み前に全シート名を検証（部分書き込みを避ける）
    for sheet_name in mapping:
        if sheet_name not in wb.sheetnames:
            raise KeyError(sheet_name)

    for sheet_name, cell_map in mapping.items():
        ws = wb[sheet_name]
        for coord, value in cell_map.items():
            ws[coord] = value

    wb.save(output_path)


def main():
    parser = argparse.ArgumentParser(description="xlsx のセルを埋める")
    parser.add_argument("--input", required=True, help="入力 xlsx パス")
    parser.add_argument("--map", required=True, help="マッピング JSON ファイルパス")
    parser.add_argument("--output", required=True, help="出力 xlsx パス")
    args = parser.parse_args()

    with open(args.map, encoding="utf-8") as f:
        mapping = json.load(f)

    fill_workbook(args.input, mapping, args.output)


if __name__ == "__main__":
    main()
