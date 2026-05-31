"""xlsx を読み込み、全非空セルの情報を構造化 dict で返す。"""
import json
import sys
from datetime import datetime, date

import openpyxl


def read_workbook(path: str) -> dict:
    """
    xlsx を2回ロード（数式取得 + キャッシュ取得）してマージし、
    全シートの非空セル情報を含む dict を返す。
    """
    wb_formula = openpyxl.load_workbook(path, data_only=False)
    wb_data = openpyxl.load_workbook(path, data_only=True)

    sheets = []
    for sheet_name in wb_formula.sheetnames:
        ws_f = wb_formula[sheet_name]
        ws_d = wb_data[sheet_name]

        cells = []
        for row in ws_f.iter_rows():
            for cell_f in row:
                if cell_f.value is None:
                    continue

                value = cell_f.value
                data_type = cell_f.data_type
                cached = None

                if data_type == "f":
                    # data_only ロードのキャッシュ値を取得
                    cell_d = ws_d[cell_f.coordinate]
                    cached = _serialize(cell_d.value)
                value = _serialize(value)

                cells.append({
                    "coord": cell_f.coordinate,
                    "row": cell_f.row,
                    "col": cell_f.column,
                    "value": value,
                    "data_type": data_type,
                    "cached": cached,
                })

        sheets.append({
            "name": sheet_name,
            "max_row": ws_f.max_row,
            "max_col": ws_f.max_column,
            "cells": cells,
        })

    return {"file": path, "sheets": sheets}


def _serialize(value):
    """datetime/date を ISO 文字列に変換する。それ以外はそのまま。"""
    if isinstance(value, datetime):
        return value.isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def main():
    if len(sys.argv) != 2:
        print("Usage: excel_reader.py <input.xlsx>", file=sys.stderr)
        sys.exit(1)
    result = read_workbook(sys.argv[1])
    print(json.dumps(result, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
